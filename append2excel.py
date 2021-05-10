"""
Created on Wed Feb 17 10:09:24 2021

@author: Audun
"Script to create dimensional time series for fatigue analysis.

   """

def append_df_to_excel(filename, df, sheet_name='Sheet1', startcol=None,    truncate_sheet=False,    **to_excel_kwargs):
    """
    #Append a DataFrame [df] to existing Excel file [filename]    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startcol : upper left cell row to dump data frame.
                 Per default (startcol=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # print('Skriver til Excel her: ' + filename)


    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startcol=startcol, **to_excel_kwargs)
    writer.save()
    
import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell

def append_to_cell(filename, variable, col=None,  row=None,  truncate_sheet=False,    **to_excel_kwargs):
    
    # Start by opening the spreadsheet and selecting the main sheet
    workbook = load_workbook(filename)
    sheet = workbook.active

    cell = xl_rowcol_to_cell(row, col)  
    print(cell)
    sheet[cell] = str(variable)
    # sheet.write(row, col, 'ISBT DEHRADUN')
    # Save the spreadsheet
    workbook.save(filename)